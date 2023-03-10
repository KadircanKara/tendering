import dash
from dash import html, dcc
from dash.dependencies import Input, Output, State
from dash.exceptions import PreventUpdate
from openpyxl.styles import Alignment
from functions import *
from objects import *
from app_pages import *

df_fiyat, df_packages, wb = load_all_sources()
df_fiyat_copy, df_packages_copy = df_fiyat.copy(), df_packages.copy()


app = dash.Dash(__name__, external_stylesheets=[dbc.themes.VAPOR], suppress_callback_exceptions=True, meta_tags=[{'name': 'viewport',
                             'content': 'width=device-width, initial-scale=1.0, maximum-scale=1.2, minimum-scale=0.5,'}]
                )

server = app.server

app.title = 'Faradai Tendering Process'

app.layout = login_page


@app.callback(

    Output('paket_table','data'),
    Input('paket_reset_button','n_clicks'),
    Input('tabs','active_tab'),
    Input('alert_paket','is_open'),
    State('alert_paket','color'),

)

def reset_paket_table(nclicks_paket_reset, tab, isopen, color):

    if nclicks_paket_reset > 0 or tab or (isopen and (color == 'success')) :
        return load_specific_packages(tab).to_dict('records')
    else :
        raise PreventUpdate

@app.callback(

    Output('fiyat_table','data'),
    Input('fiyat_reset_button','n_clicks'),
    Input('alert_fiyat','is_open'),
    State('alert_fiyat','color'),


)

def reset_fiyat_table(fiyat_reset_nclicks, isopen, color):

    if fiyat_reset_nclicks > 0 or (isopen and (color == 'success')) :
        return load_fiyat()[0].to_dict('records')
    else :
        raise PreventUpdate


@app.callback(

    Output('fiyat_table_alert','children'),
    Output('fiyat_table_alert','is_open'),
    Output('fiyat_dummy','children'),
    Input('fiyat_button','n_clicks'),
    State('fiyat_table','data'),

)

def update_fiyat(nclicks, data):

    if nclicks > 0 :

        df = pd.DataFrame.from_dict(data)
        book = xl.load_workbook('Sources/FiyatListesi.xlsx')
        ws = book.active
        column_data = [df['Iskontosuz Fiyat'], df['Iskonto']]
        current_row = 2
        current_col = 3
        for column in column_data :
            current_col += 1
            current_row = 2
            for r in range(len(column)):
                if column[r]:
                    ws.cell(row=current_row, column=current_col).value = float(column[r])
                current_row += 1
        book.save('Sources/FiyatListesi.xlsx')
        book.save('static/FiyatListesi.xlsx')
        # git_push()

        return html.Div( "Fiyat Bilgileri G??ncellendi"), True, dash.no_update

    else :
        raise PreventUpdate


@app.callback(

    Output('paket_table_alert','children'),
    Output('paket_table_alert','is_open'),
    Output('paket_dummy','children'),
    Input('paket_button','n_clicks'),
    State('tabs','active_tab'),
    State('paket_table','data'),

)

def update_paket(nclicks, tab, data):

    if nclicks > 0 :

        df = pd.DataFrame.from_dict(data)
        book = xl.load_workbook('Sources/Packages.xlsx')
        ws = book[tab]
        column_data = [df['Access'], df['Starter'], df['Standard'], df['Full-stack']]
        current_row = 2
        current_col = -1
        for column in column_data :
            current_col += 2
            current_row = 2
            for r in range(len(column)):
                print(column[r])
                if column[r]:
                    ws.cell(row=current_row, column=current_col).value = float(column[r])
                current_row += 1
        book.save('Sources/Packages.xlsx')
        book.save('static/Packages.xlsx')
        # git_push()

        return html.Div(tab + " Paketi Bilgileri G??ncellendi"), True, dash.no_update,

    else :
        raise PreventUpdate


@app.callback(

    Output('alert_fiyat','children'),
    Output('alert_fiyat','is_open'),
    Output('alert_fiyat','color'),
    Input('upload_fiyat' , 'contents'),

)

def upload_msg(fiyat_contents):

    fiyat_isopen = False
    fiyat_msg = ""
    fiyat_color = ""

    if fiyat_contents :
        fiyat_msg , fiyat_color = check_fiyat_upload(fiyat_contents)
        fiyat_isopen = True

    return fiyat_msg,fiyat_isopen,fiyat_color



@app.callback(

    Output('alert_paket','children'),
    Output('alert_paket','is_open'),
    Output('alert_paket','color'),
    Input('upload_paket' , 'contents'),

)


def upload_msg(paket_contents):

    paket_isopen = False
    paket_msg = ""
    paket_color = ""
    
    if paket_contents :
        paket_msg , paket_color = check_paket_upload(paket_contents)
        paket_isopen = True

    return paket_msg,paket_isopen,paket_color


@app.callback(

    Output('main','children'),
    Output('main','style'),
    Input('verify','n_clicks'),
    State('user','value'),
    State('passw','value'),

)

def show_page(nclicks,user,passw):

        if nclicks > 0 and user=='reengen' and passw=='rngn2021!':
            
            return (
                dbc.Container([
                    dbc.Row([
                        dbc.Col(html.Img(src='/assets/faradai_logo_nobg.png', height="60px"), width=1),
                        dbc.Col(html.H1('Faradai Tekliflendirme Modul??'), align='center', style={'color':'white'}, width=6)
                        ], align='center', justify='center', style={'align':'center','margin-top':'1rem'}
                    )
                ]),

                    html.Hr(),

                #dbc.Container([
                    html.Div(dbc.Tabs(
                    [
                        dbc.Tab(offer_page(), label="Teklif Olu??turma Sayfas??", tab_id='teklif', activeTabClassName="fw-bold fst-italic", tab_style={"margin": "auto"}, active_label_style={'backgroundColor':'transparent'}),
                        dbc.Tab(resources_page(), label="Fiyat & Paket G??ncelleme Sayfas??", tab_id='kaynak', activeTabClassName="fw-bold fst-italic", tab_style={"margin": "auto"}, active_label_style={'backgroundColor':'transparent'}),
                    ], id='page_content', active_tab='teklif', style={'width':'100%','margin':'auto'}
                ), style={'margin':'auto'})
            #])
    ) , {}
        
        else:
            raise PreventUpdate



@app.callback(
    Output("display-selected-package", "style"),
    Output("display-selected-package", "children"),
    Input("paket", "value"),
    Input("kategori", "value"),

    prevent_initial_call=True,
)
def func(paket, kat):

    if paket is not None and paket != '-' and kat is not None and kat != '-':

        df = df_packages[kat]

        df = df[[paket, '{} Cihaz'.format(paket)]]
        df.dropna(inplace=True)
        df.reset_index(inplace=True)

        df.dropna(inplace=True)

        contents = '{} - {}  : '.format(kat, paket)

        for index, row in df.iterrows():
            contents = contents + '({}) {} '.format(str(int(row[paket])), row['{} Cihaz'.format(paket)])

        contents = contents[:-1] + ''

        return {'display': 'block'}, contents
    else:
        return None, 'L??tfen kategori-paket se??imi yap??n??z.'


def cihaz_options(device, type):

    if type == 'standard':
        id = device
    elif type == 'ek':
        id = 'ek_' + device

    @app.callback(
        Output(id, "options"),

        Input("paket", "value"),
        Input("kategori", "value"),
        Input("page_content","active_tab"),
        prevent_initial_call=False,
    )

    def func(paket, kat, page_tab):
            
            if page_tab == 'teklif':

                df_fiyat = load_fiyat()[0]

                df_cihaz_list = get_adapters(df_fiyat, device)[1]
                options = [{'label': i, 'value': i} for i in df_cihaz_list['cihaz'].unique()]

                return options
            
            else :
                raise PreventUpdate


cihaz_options('Gateway', 'standard')
cihaz_options('Trifaz Analiz??r', 'standard')
cihaz_options('Ak??m Trafosu', 'standard')
cihaz_options('S??cakl??k Sens??r??', 'standard')
cihaz_options('Su Sayac??', 'standard')
cihaz_options('Ak??ll?? Klima Kontrol', 'standard')
cihaz_options('Modbus Converter', 'standard')
cihaz_options('G???? Kayna????', 'standard')
cihaz_options('Jenerat??r Kart??', 'standard')
cihaz_options('Monofaz Analiz??r', 'standard')
cihaz_options('Pulse Okuyucu', 'standard')
cihaz_options('UPS', 'standard')

cihaz_options('Gateway', 'ek')
cihaz_options('Trifaz Analiz??r', 'ek')
cihaz_options('Ak??m Trafosu', 'ek')
cihaz_options('S??cakl??k Sens??r??', 'ek')
cihaz_options('Su Sayac??', 'ek')
cihaz_options('Ak??ll?? Klima Kontrol', 'ek')
cihaz_options('Modbus Converter', 'ek')
cihaz_options('G???? Kayna????', 'ek')
cihaz_options('Jenerat??r Kart??', 'ek')
cihaz_options('Monofaz Analiz??r', 'ek')
cihaz_options('Pulse Okuyucu', 'ek')
cihaz_options('UPS', 'ek')


def cihaz_num_state(device):

    @app.callback(

        Output(device+'_num','disabled'),
        Output(device+'_num','value'),
        Input('kategori','value'),
        Input('paket','value')

    )

    def return_cihaz_num_state(kat, paket):

        if (paket is not None and paket != '-') and (kat is not None and kat != '-'):
            return True, None
        else:
            return False, dash.no_update

cihaz_num_state('Gateway')
cihaz_num_state('Trifaz Analiz??r')
cihaz_num_state('Ak??m Trafosu')
cihaz_num_state('S??cakl??k Sens??r??')
cihaz_num_state('Su Sayac??')
cihaz_num_state('Ak??ll?? Klima Kontrol')
cihaz_num_state('Modbus Converter')
cihaz_num_state('G???? Kayna????')
cihaz_num_state('Jenerat??r Kart??')
cihaz_num_state('Monofaz Analiz??r')
cihaz_num_state('Pulse Okuyucu')
cihaz_num_state('UPS')


@app.callback(
    Output("proje_adi", "valid"),
    Output("sehir_ici", "valid"),
    Output("sehir_disi", "valid"),
    Output("musteri_adi", "valid"),
    Output("teklifi_hazirlayan", "valid"),
    Output("kategori", "valid"),
    Output("paket", "valid"),

    Input("proje_adi", "value"),
    Input("sehir_ici", "value"),
    Input("sehir_disi", "value"),
    Input("musteri_adi", "value"),
    Input("teklifi_hazirlayan", "value"),
    Input("kategori", "value"),
    Input("paket", "value"),

    prevent_initial_call=True,
)
def validate1_params(proje, ic, dis, musteri, teklif, kategori, paket):
    params = [proje, ic, dis, musteri, teklif, kategori, paket]

    valid = []

    for param in params:
        if param is not None and param != '-' and param != '':
            valid.append(True)
        else:
            valid.append(False)

    return valid


@app.callback(
    Output("proje_adi", "invalid"),
    Output("sehir_ici", "invalid"),
    Output("sehir_disi", "invalid"),
    Output("musteri_adi", "invalid"),
    Output("teklifi_hazirlayan", "invalid"),

    Input("proje_adi", "value"),
    Input("sehir_ici", "value"),
    Input("sehir_disi", "value"),
    Input("musteri_adi", "value"),
    Input("teklifi_hazirlayan", "value"),

    prevent_initial_call=True,
)
def validate2_params(proje, ic, dis, musteri, teklif):
    params = [proje, ic, dis, musteri, teklif]

    invalid = []

    for param in params:
        if param is not None and param != '-' and param != '':
            invalid.append(False)
        else:
            invalid.append(True)

    return invalid


@app.callback(
    Output("submit", "disabled"),
    Output("submit", "color"),
    Input('proje_adi', 'invalid'),
    Input('musteri_adi', 'invalid'),
    Input('teklifi_hazirlayan', 'invalid'),
    Input('sehir_ici', 'invalid'),
    Input('sehir_disi', 'invalid'),
)

def enable_button(proje, musteri, teklif, ic, dis):
    params = [proje, musteri, teklif, ic, dis]

    if True in params:
        return True, 'danger'
    else:
        return False, 'secondary'


@app.callback(

    Output("Output-Status", "children"),
    Output("Output-Status", "dismissable"),
    Output("Output-Status", "is_open"),
    Output("Output-Status", "duration"),
    Output('download-file-xlsx', 'data'),

    Input("submit", "n_clicks"),
    State("proje_adi", "value"),
    State("musteri_adi", "value"),
    State("teklifi_hazirlayan", "value"),
    State("sehir_ici", "value"),
    State("sehir_disi", "value"),
    State("kategori", "value"),
    State("paket", "value"),

    State("Gateway", "value"),
    State("Gateway_num", "value"),
    State("ek_Gateway", "value"),
    State("ek_Gateway_ic", "value"),
    State("ek_Gateway_dis", "value"),

    State("Trifaz Analiz??r", "value"),
    State("Trifaz Analiz??r_num", "value"),
    State("ek_Trifaz Analiz??r", "value"),
    State("ek_Trifaz Analiz??r_ic", "value"),
    State("ek_Trifaz Analiz??r_dis", "value"),

    State("Ak??m Trafosu", "value"),
    State("Ak??m Trafosu_num", "value"),
    State("ek_Ak??m Trafosu", "value"),
    State("ek_Ak??m Trafosu_ic", "value"),
    State("ek_Ak??m Trafosu_dis", "value"),

    State("S??cakl??k Sens??r??", "value"),
    State("S??cakl??k Sens??r??_num", "value"),
    State("ek_S??cakl??k Sens??r??", "value"),
    State("ek_S??cakl??k Sens??r??_ic", "value"),
    State("ek_S??cakl??k Sens??r??_dis", "value"),

    State("Su Sayac??", "value"),
    State("Su Sayac??_num", "value"),
    State("ek_Su Sayac??", "value"),
    State("ek_Su Sayac??_ic", "value"),
    State("ek_Su Sayac??_dis", "value"),

    State("Ak??ll?? Klima Kontrol", "value"),
    State("Ak??ll?? Klima Kontrol_num", "value"),
    State("ek_Ak??ll?? Klima Kontrol", "value"),
    State("ek_Ak??ll?? Klima Kontrol_ic", "value"),
    State("ek_Ak??ll?? Klima Kontrol_dis", "value"),

    State("Modbus Converter", "value"),
    State("Modbus Converter_num", "value"),
    State("ek_Modbus Converter", "value"),
    State("ek_Modbus Converter_ic", "value"),
    State("ek_Modbus Converter_dis", "value"),

    State("G???? Kayna????", "value"),
    State("G???? Kayna????_num", "value"),
    State("ek_G???? Kayna????", "value"),
    State("ek_G???? Kayna????_ic", "value"),
    State("ek_G???? Kayna????_dis", "value"),

    State("Jenerat??r Kart??", "value"),
    State("Jenerat??r Kart??_num", "value"),
    State("ek_Jenerat??r Kart??", "value"),
    State("ek_Jenerat??r Kart??_ic", "value"),
    State("ek_Jenerat??r Kart??_dis", "value"),

    State("Monofaz Analiz??r", "value"),
    State("Monofaz Analiz??r_num", "value"),
    State("ek_Monofaz Analiz??r", "value"),
    State("ek_Monofaz Analiz??r_ic", "value"),
    State("ek_Monofaz Analiz??r_dis", "value"),

    State("Pulse Okuyucu", "value"),
    State("Pulse Okuyucu_num", "value"),
    State("ek_Pulse Okuyucu", "value"),
    State("ek_Pulse Okuyucu_ic", "value"),
    State("ek_Pulse Okuyucu_dis", "value"),

    State("UPS", "value"),
    State("UPS_num", "value"),
    State("ek_UPS", "value"),
    State("ek_UPS_ic", "value"),
    State("ek_UPS_dis", "value"),

    prevent_initial_call=True

)
def write_to_excel(nclicks, proje, musteri, teklif, ic, dis, kategori, paket, gateway, gateway_num, ek_gateway, gateway_ic,
                   gateway_dis,
                   trifaz, trifaz_num,  ek_trifaz, trifaz_ic, trifaz_dis, akim, akim_num, ek_akim, akim_ic, akim_dis, sicaklik, sicaklik_num, ek_sicaklik,
                   sicaklik_ic, sicaklik_dis,
                   su, su_num, ek_su, su_ic, su_dis, klima, klima_num, ek_klima, klima_ic, klima_dis, modbus, modbus_num, ek_modbus, modbus_ic,
                   modbus_dis,
                   guc, guc_num, ek_guc, guc_ic, guc_dis, jenerator, jenerator_num, ek_jenerator, jenerator_ic, jenerator_dis, monofaz, monofaz_num,
                   ek_monofaz, monofaz_ic, monofaz_dis,
                   pulse, pulse_num, ek_pulse, pulse_ic, pulse_dis, ups, ups_num, ek_ups, ups_ic, ups_dis):

    
    df_fiyat, df_packages, wb = load_all_sources()

    if ic is None : ic = 0
    if dis is None : dis = 0
    wb['Teklif??al????mas??']['C6'] = ic + dis

    donanimlar = [gateway, trifaz, akim, sicaklik, su, klima, modbus, guc, jenerator, monofaz, pulse, ups]
    donanimlar_adet = [gateway_num, trifaz_num, akim_num, sicaklik_num, su_num, klima_num, modbus_num, guc_num, jenerator_num, monofaz_num, pulse_num, ups_num]
    cihazlar = ['Gateway', 'Trifaz', 'Ak??m Trafosu', 'S??cakl??k Sens??r??', 'Su Sayac??', 'Ak??ll?? Klima Kontrol',
                'Modbus Converter',
                'G???? Kayna????', 'Jenerat??r Kart??', 'Monofaz Analiz??r', 'Pulse Okuyucu', 'UPS']

    ek_donanimlar = [ek_gateway, ek_trifaz, ek_akim, ek_sicaklik, ek_su, ek_klima, ek_modbus, ek_guc,
                     ek_jenerator, ek_monofaz, ek_pulse, ek_ups]


    ek_donanim_ic = [gateway_ic, trifaz_ic, akim_ic, sicaklik_ic, su_ic, klima_ic, modbus_ic, guc_ic, jenerator_ic,
                     monofaz_ic, pulse_ic, ups_ic]
    ek_donanim_dis = [gateway_dis, trifaz_dis, akim_dis, sicaklik_dis, su_dis, klima_dis, modbus_dis, guc_dis,
                      jenerator_dis, monofaz_dis, pulse_dis, ups_dis]

    ek_donanim_ic_toplam, ek_donanim_dis_toplam = 0, 0

    for i in range(len(ek_donanimlar)):
        if ek_donanimlar[i] is not None and ek_donanimlar[i] != '-':
            if ek_donanim_ic[i] is not None:
                ek_donanim_ic_toplam += ek_donanim_ic[i]
            if ek_donanim_dis[i] is not None:
                ek_donanim_dis_toplam += ek_donanim_dis[i]


# -----------------------------------------------   START EXCEL OPERATIONS  -----------------------------------------

    ws = wb['Yat??r??mMaliyeti']

    ws['G3'] = tcmb_data()
    ws['G4'] = tcmb_data('EUR')
    ws['G5'] = tcmb_data('GBP')

    ws.cell(row=4, column=3).value = musteri
    ws.cell(row=5, column=3).value = proje
    ws.cell(row=6, column=3).value = teklif

    ws['C8'] = paket

    ws['H9'] = '??ehir ????i : ' + str(ic)
    ws['I9'] = '??ehir D?????? : ' + str(dis)

    for i in range(len(donanimlar)):

        print(ek_donanimlar[i],ek_donanim_ic[i])

        if donanimlar[i] is not None and donanimlar[i] != '-':

            ws['B{}'.format(i + 15)] = donanimlar[i]

            cihaz = cihazlar[i]

            if (paket is not None and paket != '-') and (kategori is not None and kategori != '-'):
                df = df_packages[kategori]
                df = df[[paket, '{} Cihaz'.format(paket)]]
                df.dropna(inplace=True)
                df.reset_index(inplace=True)
                df.dropna(inplace=True)
                df_kurulumsuz = df[df['{} Cihaz'.format(paket)].str.contains("Kurulum") == False]
                ws['E{}'.format(i + 15)] = df_kurulumsuz.loc[df_kurulumsuz['{} Cihaz'.format(paket)] == cihaz][paket].sum()
            else:
                if donanimlar_adet[i] is not None:
                    ws['E{}'.format(i + 15)] = donanimlar_adet[i]
                else:
                    ws['E{}'.format(i + 15)] = 0

            ws['G{}'.format(i + 15)] = df_fiyat.loc[df_fiyat['Adapt??r'] == donanimlar[i]]['Iskontosuz Fiyat'].values[0] * (1 - df_fiyat.loc[df_fiyat['Adapt??r'] == donanimlar[i]]['Iskonto'].values[0])
            ws['F{}'.format(i + 15)] = df_fiyat.loc[df_fiyat['Adapt??r'] == donanimlar[i]]['Para Birimi'].values[0]

            ws['E{}'.format(i+15)].alignment = Alignment(horizontal='center')
            ws['F{}'.format(i+15)].alignment = Alignment(horizontal='center')
            ws['G{}'.format(i+15)].alignment = Alignment(horizontal='center')
            ws['H{}'.format(i+15)].alignment = Alignment(horizontal='center')

        else:
            ws['B{}'.format(i + 15)], ws['E{}'.format(i + 15)], ws['G{}'.format(i + 15)], ws[
                'F{}'.format(i + 15)] = "", "", "", ""

        if (ek_donanimlar[i] is not None and ek_donanimlar[i] != '-') and (
                ek_donanim_ic[i] is not None or ek_donanim_dis[i] is not None) and (
                int(bool(ek_donanim_ic[i])) + int(bool(ek_donanim_dis[i]))) != 0:

            ws['B{}'.format(i + 73)] = ek_donanimlar[i]
            cihaz = cihazlar[i]
            ic_dis_donanim_sayisi = 0
            ek_donanim_aciklama = ''
            if ek_donanim_ic[i]:
                ic_dis_donanim_sayisi += ek_donanim_ic[i]
                ek_donanim_aciklama += ('???? : ' + str(ek_donanim_ic[i]) + ' ')
            if ek_donanim_dis[i]:
                ic_dis_donanim_sayisi += ek_donanim_dis[i]
                ek_donanim_aciklama += ('D???? : ' + str(ek_donanim_dis[i]))
            ws['E{}'.format(i + 73)] = ic_dis_donanim_sayisi
            ws['I{}'.format(i + 73)] = ek_donanim_aciklama
            ws['G{}'.format(i + 73)] = df_fiyat.loc[df_fiyat['Adapt??r'] == ek_donanimlar[i]]['Iskontosuz Fiyat'].values[0] * (1 - df_fiyat.loc[df_fiyat['Adapt??r'] == ek_donanimlar[i]]['Iskonto'].values[0])
            ws['F{}'.format(i + 73)] = df_fiyat.loc[df_fiyat['Adapt??r'] == ek_donanimlar[i]]['Para Birimi'].values[0]

        else:
            ws['B{}'.format(i + 73)], ws['E{}'.format(i + 73)], ws['I{}'.format(i + 73)], ws['G{}'.format(i + 73)], ws[
                'F{}'.format(i + 73)] = "", "", "", "", ""

    if ic is not None or ic != 0:
        ws['B33'] = 'Paket Kurulum (??ehir ????i)'
        ws['G33'] = df_fiyat.loc[df_fiyat['Adapt??r'] == 'Paket Kurulum (??ehir i??i)']['Iskontosuz Fiyat'].values[0] * (1 - df_fiyat.loc[df_fiyat['Adapt??r'] == 'Paket Kurulum (??ehir i??i)']['Iskonto'].values[0])
        ws['E33'] = ic
        ws['F33'] = df_fiyat.loc[df_fiyat['Adapt??r'] == 'Paket Kurulum (??ehir i??i)']['Para Birimi'].values[0]
#         ws['B35'] = 'Cihaz Ba???? Kurulum (??ehir i??i)'
#         ws['G35'] = df_fiyat.loc[df_fiyat['Adapt??r'] == 'Cihaz Ba???? Kurulum (??ehir i??i)']['Iskontosuz Fiyat'].values[0] * (1 - df_fiyat.loc[df_fiyat['Adapt??r'] == 'Cihaz Ba???? Kurulum (??ehir i??i)']['Iskonto'].values[0])
#         ws['E35'] = "={}+(SUM(E17:E26)*{})".format(ek_donanim_ic_toplam, ic)
#         ws['F35'] = df_fiyat.loc[df_fiyat['Adapt??r'] == 'Cihaz Ba???? Kurulum (??ehir i??i)']['Para Birimi'].values[0]

    if dis is not None or dis != 0:
        ws['B34'] = 'Paket Kurulum (??ehir D??????)'
        ws['G34'] = df_fiyat.loc[df_fiyat['Adapt??r'] == 'Paket Kurulum (??ehir d??????)']['Iskontosuz Fiyat'].values[0] * (1 - df_fiyat.loc[df_fiyat['Adapt??r'] == 'Paket Kurulum (??ehir d??????)']['Iskonto'].values[0])
        ws['E34'] = dis
        ws['F34'] = df_fiyat.loc[df_fiyat['Adapt??r'] == 'Paket Kurulum (??ehir d??????)']['Para Birimi'].values[0]
#         ws['B36'] = 'Cihaz Ba???? Kurulum (??ehir d??????)'
#         ws['G36'] = df_fiyat.loc[df_fiyat['Adapt??r'] == 'Cihaz Ba???? Kurulum (??ehir d??????)']['Iskontosuz Fiyat'].values[0] * (1 - df_fiyat.loc[df_fiyat['Adapt??r'] == 'Cihaz Ba???? Kurulum (??ehir d??????)']['Iskonto'].values[0])
#         ws['E36'] = "={}+(SUM(E17:E26)*{})".format(ek_donanim_dis_toplam, dis)
#         ws['F36'] = df_fiyat.loc[df_fiyat['Adapt??r'] == 'Cihaz Ba???? Kurulum (??ehir d??????)']['Para Birimi'].values[0]

    ws['F33'].alignment = Alignment(horizontal='center')
    ws['F34'].alignment = Alignment(horizontal='center')
    ws['F35'].alignment = Alignment(horizontal='center')

    name_of_file = musteri + " - " + proje
    wb.save('Outputs/' + name_of_file + '.xlsx')

    return html.Div(name_of_file + " adl?? dosyan??z haz??r.",style={'text-align':'center','height':'auto'}), True, True, 4000, dcc.send_file('Outputs/' + name_of_file + '.xlsx')


@app.callback(

    Output('Gateway', 'disabled'),
    Output('Trifaz Analiz??r', 'disabled'),
    Output('Ak??m Trafosu', 'disabled'),
    Output('S??cakl??k Sens??r??', 'disabled'),
    Output('Su Sayac??', 'disabled'),
    Output('Ak??ll?? Klima Kontrol', 'disabled'),
    Output('Modbus Converter', 'disabled'),
    Output('G???? Kayna????', 'disabled'),
    Output('Jenerat??r Kart??', 'disabled'),
    Output('Monofaz Analiz??r', 'disabled'),
    Output('Pulse Okuyucu', 'disabled'),
    Output('UPS', 'disabled'),

    Output('Gateway', 'value'),
    Output('Trifaz Analiz??r', 'value'),
    Output('Ak??m Trafosu', 'value'),
    Output('S??cakl??k Sens??r??', 'value'),
    Output('Su Sayac??', 'value'),
    Output('Ak??ll?? Klima Kontrol', 'value'),
    Output('Modbus Converter', 'value'),
    Output('G???? Kayna????', 'value'),
    Output('Jenerat??r Kart??', 'value'),
    Output('Monofaz Analiz??r', 'value'),
    Output('Pulse Okuyucu', 'value'),
    Output('UPS', 'value'),

    Output('ek_Gateway', 'value'),
    Output('ek_Trifaz Analiz??r', 'value'),
    Output('ek_Ak??m Trafosu', 'value'),
    Output('ek_S??cakl??k Sens??r??', 'value'),
    Output('ek_Su Sayac??', 'value'),
    Output('ek_Ak??ll?? Klima Kontrol', 'value'),
    Output('ek_Modbus Converter', 'value'),
    Output('ek_G???? Kayna????', 'value'),
    Output('ek_Jenerat??r Kart??', 'value'),
    Output('ek_Monofaz Analiz??r', 'value'),
    Output('ek_Pulse Okuyucu', 'value'),
    Output('ek_UPS', 'value'),

    Input("kategori", "value"),
    Input("paket", "value"),
    Input('Gateway', 'value'),
    Input('Trifaz Analiz??r', 'value'),
    Input('Ak??m Trafosu', 'value'),
    Input('S??cakl??k Sens??r??', 'value'),
    Input('Su Sayac??', 'value'),
    Input('Ak??ll?? Klima Kontrol', 'value'),
    Input('Modbus Converter', 'value'),
    Input('G???? Kayna????', 'value'),
    Input('Jenerat??r Kart??', 'value'),
    Input('Monofaz Analiz??r', 'value'),
    Input('Pulse Okuyucu', 'value'),
    Input('UPS', 'value'),

    Input('page_content','active_tab'),

    prevent_initial_call=False

)
def required_devices_border(kategori, paket, gateway, trifaz, akim, sicaklik, su, klima, modbus, guc, jenerator,
                            monofaz, pulse, ups, page_tab):

    all_inputs = [gateway, trifaz, akim, sicaklik, su, klima, modbus, guc, jenerator, monofaz, pulse, ups]

    DISABLED = [True] * len(all_inputs)
    VALUE = all_inputs

    DISABLED = [True] * len(all_inputs)
    VALUE = all_inputs

    if page_tab == 'kaynak':
        return [dash.no_update]*len(all_inputs) + [None]*(2*len(all_inputs))
    
    else :

        if (kategori is not None and kategori != '-') and (paket is not None and paket != '-'):

            df = df_packages[kategori]

            df = df[[paket, '{} Cihaz'.format(paket)]]
            df.dropna(inplace=True)
            df.reset_index(drop=True)

            df.dropna(inplace=True)

            tum_cihazlar = ['Gateway', 'Trifaz', 'Ak??m Trafosu', 'S??cakl??k Sens??r??', 'Su Sayac??', 'Ak??ll?? Klima Kontrol',
                            'Modbus Converter',
                            'G???? Kayna????', 'Jenerat??r Kart??', 'Monofaz Analiz??r', 'Pulse Okuyucu', 'UPS']
            
            if kategori == '-' or paket == '-':
                gereken_cihazlar = tum_cihazlar
            else:
                gereken_cihazlar = df.loc[df['{} Cihaz'.format(paket)] != "Kurulum"]['{} Cihaz'.format(paket)].unique().tolist()

            girilen_cihazlar = [gateway, trifaz, akim, sicaklik, su, klima, modbus, guc, jenerator, monofaz, pulse, ups]

            for i in range(len(tum_cihazlar)):
                if tum_cihazlar[i] in gereken_cihazlar:
                    DISABLED[i] = False
                else:
                    VALUE[i] = None

            return DISABLED + VALUE + [dash.no_update]*len(all_inputs)


        else:

            DISABLED = [False] * len(all_inputs)
            return DISABLED + VALUE + [dash.no_update]*len(all_inputs)


if __name__ == "__main__":
    app.run_server(debug=True)
