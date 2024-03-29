import numpy as np
import base64
import io
import os
import openpyxl as xl
import pandas as pd
from tradingview_ta import TA_Handler, Interval
from dash import  html
import dash_bootstrap_components as dbc
from objects import *
import os

def load_specific_packages(packages:list or str):

    if isinstance(packages,list):

        df_packages = {}

        for package in packages:
            df_packages.update({package:pd.read_excel("Sources/Packages.xlsx", sheet_name=package)})

        for df in df_packages.values():
            df.rename(columns={'Unnamed: 1':'Access Cihaz' , 'Unnamed: 3':'Starter Cihaz' , 'Unnamed: 5':'Standard Cihaz' , 'Unnamed: 7':'Full-stack Cihaz'},inplace=True)

        return df_packages

    elif isinstance(packages,str):
        df = pd.read_excel("Sources/Packages.xlsx", sheet_name=packages)
        df.rename(columns={'Unnamed: 1':'Access Cihaz' , 'Unnamed: 3':'Starter Cihaz' , 'Unnamed: 5':'Standard Cihaz' , 'Unnamed: 7':'Full-stack Cihaz'},inplace=True)
        return df

def load_fiyat():
    df_fiyat = pd.read_excel("Sources/FiyatListesi.xlsx")
    df_fiyat.columns.str.match("Unnamed")
    df_fiyat.drop(df_fiyat.loc[:,df_fiyat.columns.str.match("Unnamed")], axis=1, inplace=True)
    columns = list(df_fiyat.columns)
    return df_fiyat , columns

def load_all_sources():

    df_fiyat = pd.read_excel("Sources/FiyatListesi.xlsx")
    df_retail = pd.read_excel("Sources/Packages.xlsx" , sheet_name='Retail')
    df_Banking = pd.read_excel("Sources/Packages.xlsx", sheet_name='Banking')
    df_Hospital = pd.read_excel("Sources/Packages.xlsx", sheet_name='Hospital')
    df_Supermarkets = pd.read_excel("Sources/Packages.xlsx", sheet_name='Supermarkets')
    df_Industry = pd.read_excel("Sources/Packages.xlsx", sheet_name='Industry')
    wb = xl.load_workbook("Sources/Source.xlsx")

    df_fiyat.columns.str.match("Unnamed")
    df_fiyat.drop(df_fiyat.loc[:,df_fiyat.columns.str.match("Unnamed")], axis=1, inplace=True)

    df_packages = {'Retail':df_retail, 'Banking':df_Banking, 'Hospital':df_Hospital, 'Supermarkets':df_Supermarkets, 'Industry':df_Industry}

    for df in df_packages.values():
        df.rename(columns={'Unnamed: 1':'Access Cihaz' , 'Unnamed: 3':'Starter Cihaz' , 'Unnamed: 5':'Standard Cihaz' , 'Unnamed: 7':'Full-stack Cihaz'},inplace=True)

    return df_fiyat, df_packages, wb

def check(file_name):
    return os.path.exists(file_name)

def tcmb_data(kur="USD"):
    c_exchange = TA_Handler(
        symbol="{}TRY".format(kur),
        screener="forex",
        exchange="FX_IDC",
        interval=Interval.INTERVAL_1_DAY
    )
    a = round(c_exchange.get_analysis().indicators["high"] + 0.2, 2)
    return str(a).replace('.', ',')

def get_adapters(df, cihaz):
    #df = df.loc[df['Cihaz Türü'] == cihaz].reset_index(drop=True)
    df = df[df['Cihaz Türü'].str.contains(cihaz)].reset_index(drop=True)

    # df_cihaz = pd.DataFrame(df.loc[df['Cihaz Türü'] == cihaz].reset_index(drop=True)['Adaptör'])
    df_cihaz = pd.DataFrame(df[df['Cihaz Türü'].str.contains(cihaz)].reset_index(drop=True)['Adaptör'])
    cihaz_list = ['-'] + df_cihaz['Adaptör'].tolist()
    # cihaz_list.append('-')

    # df_birim = pd.DataFrame(df.loc[df['Cihaz Türü'] == cihaz].reset_index(drop=True)['Para Birimi'])
    df_birim = pd.DataFrame(df[df['Cihaz Türü'].str.contains(cihaz)].reset_index(drop=True)['Para Birimi'])
    # df_fiyat = pd.DataFrame(df.loc[df['Cihaz Türü'] == cihaz].reset_index(drop=True)['Iskontosuz Fiyat'] * (1 - df.loc[df['Cihaz Türü'] == cihaz].reset_index(drop=True)['Iskonto']))
    df_fiyat = pd.DataFrame(df[df['Cihaz Türü'].str.contains(cihaz)].reset_index(drop=True)['Iskontosuz Fiyat'] * (1 - df[df['Cihaz Türü'].str.contains(cihaz)].reset_index(drop=True)['Iskonto']))

    df = pd.concat([df_cihaz, df_fiyat, df_birim], axis=1)

    df_cihaz_list = pd.DataFrame({
        'cihaz': cihaz_list
    })

    return df, df_cihaz_list

def lower_string_list(stringList):
    out = []
    for string in stringList :
       out.append(string.lower())
    return out

def every_first_letter_uppercase(string):

    while string[0] == ' ' or string[-1] == ' ':
        if string[0] == ' ':
            string = string[1:] # Delete initial whitespaces if there are any
        if string[-1] == ' ':
            string = string[0:-1] # Delete last whitespaces if there are any
    
    string = string[0].upper() + string[1:] # First letter uppercase

    for i in range(len(string)):
        if string[i] == ' ':
            string = string[0:i+1] + string[i+1].upper() + string[i+2:]

    return string

def check_fiyat_upload(contents, title):

    content_type, content_string = contents.split(',')
    decoded = base64.b64decode(content_string)
    wb = xl.load_workbook(io.BytesIO(decoded))
    sheets = wb.sheetnames

    print(sheets)

    color = 'success'

    if title == 'Faradai Tekliflendirme Modulü':
        success_msg = 'Fiyat Dosyası Yüklendi'
    else :
        success_msg = 'Price File Uploaded'

    df_fiyat = pd.read_excel(io.BytesIO(decoded) , sheet_name=sheets[0])

    headers , fiyat_headers = list(df_fiyat.columns) , ['Cihaz Türü','Adaptör','Para Birimi','Iskontosuz Fiyat','Iskonto']
    headers_lowercase , fiyat_headers_lowercase = lower_string_list(headers) , lower_string_list(fiyat_headers)
    print(headers)

    faulty_headers = []
    wrong_data = []
    wrong_rate_flag = 0
    wrong_rate = []
    wrong_number = []

    string_data = headers[0:2]
    rate_data = headers[2]
    number_data = headers[3:]

    msg = ''
    error_msgs = []

    for header in fiyat_headers_lowercase :
        if header not in headers_lowercase :
            faulty_headers.append(header)

    if faulty_headers:

        for i in range(len(faulty_headers)):
            msg += '"' + every_first_letter_uppercase(faulty_headers[i]) + '"' + ' '
        msg = msg[0:-1] + '. '
        if title == 'Faradai Tekliflendirme Modulü':
            return 'Bazı sütun(lar) bulunamadı : ' + msg , 'danger'
        else :
            return 'Some column(s) not found : ' + msg , 'danger'

    else:


        for header in headers:

            print(header.lower() , number_data)
            
            data = df_fiyat[header]
            if header.lower() in lower_string_list(string_data + [rate_data]) :
                for i,v in enumerate(data):
                    if isinstance(v,int) or isinstance(v,float) :
                        wrong_data.append(header)

            if header.lower() == rate_data.lower():
                for i,v in enumerate(data):
                    if v.upper() not in ['EUR','TL','USD']:
                        wrong_rate_flag = 1
                        wrong_rate.append(v.upper())

            if header.lower() in lower_string_list(number_data) and 'unnamed' not in header.lower() :
                print(data.dtypes)
                if data.dtypes != np.float64 and data.dtypes != np.int64 :
                    wrong_number.append(header)

        
        if wrong_data :
            if title == 'Faradai Tekliflendirme Modulü':
                msg += 'Bazı sütun(lar)da yanlış veri tipi tespit edilmiştir : '
            else :
                msg += 'Incorrect data type found on some column(s) : '
            for header in wrong_data :
                msg += f'"{header}" , '
            msg = msg[0:-3] + '.'
        
        elif wrong_rate_flag :
            if title == 'Faradai Tekliflendirme Modulü':
                msg += '"Para Birimi" sütununda yanlış para birim(leri) tespit edilmiştir : '
            else :
                msg += 'Incorrect currency or currencies found under "Currency" column : '
            for rate in list(set(wrong_rate)) :
                msg += f'"{rate}" , '
            msg = msg[0:-3] + '.'

        if wrong_number:
            if title == 'Faradai Tekliflendirme Modulü':
                msg += 'Bazı fiyat sütun(lar)ında yanlış veri tipi tespit edilmiştir : '
            else :
                msg += 'Incorrect data type found under price column : '
            for header in wrong_number :
                msg += f'"{header}" , '
            msg = msg[0:-3] + '.'


        if wrong_data or wrong_rate_flag or wrong_number :
            msg_list = msg.split('.')
            for i,v in enumerate(msg_list):
                msg_list[i] = html.Div(v)
            color = 'danger'
            return msg_list,color


        else :

            wb.save('Sources/FiyatListesi.xlsx')
            wb.save('static/FiyatListesi.xlsx')
            color = 'success'
            return success_msg , color

def check_paket_upload(contents, title):

    content_type, content_string = contents.split(',')
    decoded = base64.b64decode(content_string)
    wb = xl.load_workbook(io.BytesIO(decoded))
    sheets = wb.sheetnames

    if len(sheets) >= len(packages) :

        color = 'success'
        if title == 'Faradai Tekliflendirme Modulü':
            success_msg = 'Paket Dosyası Yüklendi'
        else :
            success_msg = 'Package File Uploaded'
        missing_sheets = []

        for i in packages:
            if i.lower() not in lower_string_list(sheets):
                missing_sheets.append(i)

        for sheet in sheets:
            ss_sheet = wb[sheet]
            ss_sheet.title = every_first_letter_uppercase(sheet)

        sheets = wb.sheetnames
        for sheet in sheets :
            if sheet not in ['Retail','Banking','Hospital','Supermarkets','Industry'] :
                sheets.pop(sheets.index(sheet))


        dfs = {}
        headers = {}
        string_data = {}
        number_data = {}
        faulty_headers = []
        wrong_data = []
        wrong_number = []

        if title == 'Faradai Tekliflendirme Modulü':
            msg = 'Paket Dosyası Yüklendi'
        else :
            msg = 'Package File Uploaded'
        color = 'success'
        error_msgs = []

        for sheet in sheets :
            dfs.update( {sheet:pd.read_excel(io.BytesIO(decoded) , sheet_name=sheet)} )
            dfs[sheet] = dfs[sheet].rename(columns={'Unnamed: 1':'Access Cihaz' , 'Unnamed: 3':'Starter Cihaz' , 'Unnamed: 5':'Standard Cihaz' , 'Unnamed: 7':'Full-stack Cihaz'})
            dfs[sheet] = dfs[sheet].dropna()
            headers.update( {sheet:list(dfs[sheet].columns)} )
            string_data.update( {sheet:[dfs[sheet][column] for column in headers[sheet][1::2]]} )
            number_data.update( {sheet:[dfs[sheet][column] for column in headers[sheet][0::2]]} )


        paket_headers = ['Access','Access Cihaz','Starter','Starter Cihaz','Standard','Standard Cihaz','Full-stack','Full-stack Cihaz']

        headers_lowercase = headers.copy()
        for header in headers_lowercase.values() :
            header = lower_string_list(header)

        paket_headers_lowercase = lower_string_list(paket_headers)

        for df in dfs.values() :

            for comb in dfs.items():
                if df.equals(comb[1]):
                    category = comb[0]
                    break


            if lower_string_list(list(df.columns)) != paket_headers_lowercase:
                faulty_headers.append(category)

            for column_data in number_data[category] :
                if column_data.dtypes == np.dtype(np.object_):
                    wrong_data.append(category)

                for value in column_data :
                    if value <= 0 :
                        wrong_number.append(category)


            for column_data in string_data[category] :
                for value in column_data :
                    if not isinstance(value,str) and category not in wrong_data :
                        wrong_data.append(category)
            
        if wrong_number:

            msg = ''
            color = 'danger'

            for i in wrong_number :
                msg += i + ', '
            if len(wrong_number) > 1 :
                if title == 'Faradai Tekliflendirme Modulü':
                    msg += "paketlerinde hatalı cihaz sayı(lar)ı tespit edildi (0 veya 0'dan küçük). Lütfen kontrol ediniz."
                else :
                    msg += "packages include some incorrect device number(s) (0 or less than 0). Please check."
            else :
                if title == 'Faradai Tekliflendirme Modulü':
                    msg += "paketinde hatalı cihaz sayı(lar)ı tespit edildi (0 veya 0'dan küçük). Lütfen kontrol ediniz."
                else :
                    msg += "package includes some incorrect device number(s) (0 or less than 0). Please check."


        elif wrong_data:

            msg = ''
            color = 'danger'

            for i in wrong_data :
                msg += i + ', '
            if len(wrong_data) > 1 :
                if title == 'Faradai Tekliflendirme Modulü':
                    msg += 'paketlerinde hatalı veri tip(ler)i tespit edildi. Lütfen kontrol ediniz.'
                else :
                    msg += 'packages include some incorrect data type(s). Please check.'
            else :
                if title == 'Faradai Tekliflendirme Modulü':
                    msg += 'paketinde hatalı veri tip(ler)i tespit edildi. Lütfen kontrol ediniz.'
                else :
                    msg += 'package includes some incorrect data type(s). Please check.'


        elif faulty_headers :

            msg = ''
            color = 'danger'
            
            for i in faulty_headers :
                msg += i + ', '
            if len(faulty_headers) > 1 :
                if title == 'Faradai Tekliflendirme Modulü':
                    msg += 'paketlerinin başlıkları hatalı. Lütfen kontrol ediniz.'
                else :
                    msg += 'packages have incorrect headers. Please check.'
            else :
                if title == 'Faradai Tekliflendirme Modulü':
                    msg += 'paketinin başlıkları hatalı. Lütfen kontrol ediniz.'
                else :
                    msg += 'package has incorrect headers. Please check.'

        
        else :

            wb.save('Sources/Packages.xlsx')
            wb.save('static/Packages.xlsx')

        return msg,color

    else :
        if title == 'Faradai Tekliflendirme Modulü':
            return 'Yüklediğiniz exceldeki sayfa sayısı hatalı. Lütfen kontrol ediniz.', 'danger'
        else :
            return 'Uploaded excel file has an incorrect number of sheets. Please check.', 'danger'


def cihaz_row(cihaz, lang):

    if lang != 'tr':
        cihaz_tr = device_group_names_eng_to_tr[cihaz]

    if lang == 'tr':

        return (

            html.Div(children=[

                dbc.Row(children=[
            
                    dbc.Col(
                        html.Label(cihaz, style=input_style), align='center', width=2
                    ),

                    dbc.Col(
                        dbc.Select(id=cihaz, placeholder="", disabled=False), align='center', width=2
                    ),

                    dbc.Col(
                        dbc.Input(id='{}_num'.format(cihaz), type='number', placeholder='#', min=0, step=1, style=small_input_style), align='center', width=1
                    ),

                    dbc.Col(
                        html.Label("Ek {}".format(cihaz), style=input_style), align='center', width={'size':2,'offset':1}
                    ),

                    dbc.Col(
                        dbc.Select(id='ek_{}'.format(cihaz), placeholder=""), align='center', width=2
                    ),

                    dbc.Col(
                        dbc.Input(id='ek_{}_ic'.format(cihaz), type='number', placeholder='İç', min=0, step=1 , style=small_input_style), align='center', width=1
                    ),

                    dbc.Col(
                        dbc.Input(id='ek_{}_dis'.format(cihaz), type='number', placeholder='Dış', min=0, step=1 , style=small_input_style), align='center', width=1
                    ),
            
                ]),

                html.Br(),

            ])
        )
    
    else :

        return (

            html.Div(children=[

                dbc.Row(children=[
            
                    dbc.Col(
                        html.Label(cihaz, style=input_style), align='center', width=2
                    ),

                    dbc.Col(
                        dbc.Select(id= cihaz_tr, placeholder="", disabled=False), align='center', width=2
                    ),

                    dbc.Col(
                        dbc.Input(id='{}_num'.format(cihaz_tr), type='number', placeholder='#', min=0, step=1, style=small_input_style), align='center', width=1
                    ),

                    dbc.Col(
                        html.Label("Extra {}".format(cihaz), style=input_style), align='center', width={'size':2,'offset':1}
                    ),

                    dbc.Col(
                        dbc.Select(id='ek_{}'.format(cihaz_tr), placeholder=""), align='center', width=2
                    ),

                    dbc.Col(
                        dbc.Input(id='ek_{}_ic'.format(cihaz_tr), type='number', placeholder='Inter', min=0, step=1 , style=small_input_style), align='center', width=1
                    ),

                    dbc.Col(
                        dbc.Input(id='ek_{}_dis'.format(cihaz_tr), type='number', placeholder='Intra', min=0, step=1 , style=small_input_style), align='center', width=1
                    ),
            
                ]),

                html.Br(),

            ])
        )
